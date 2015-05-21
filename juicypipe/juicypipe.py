#!/usr/bin/env python
#-*- coding: utf-8 -*
"""

"""
from pipe import Pipe
import warnings
import collections
import sys
import os
import glob
import bz2
import gzip
import itertools
import sqlite3
try:
    import openpyxl
except ImportError:
    warnings.warn("You cant output into Excel. Please `pip install openpyxl`.")

from .io import *

##################################################
## Pipe関数群
##################################################
@Pipe
def vertical_map(iterable, colname, func, newcolname=None):
    """特定のカラムの全行に関数を適用する(マップ演算)
    newcolnameを指定すると新しい名前のカラムをつくる
    """
    if newcolname is None: newcolname=colname
    for item in iterable:
        item[newcolname] = func(item[colname])
        yield item

@Pipe
def pipe_round(iterable, colname, precision=0):
    """丸め
    round標準関数があるのでroundという名前は使えない"""
    for item in iterable:
        item[colname] = round(item[colname], precision)
        yield item

@Pipe
def topivot(iterable, keycols, sumcols, delim="\t", to_count=False):
    """カラムをグループ化してピボット形式にする"""
    grouped_data = {}
    def line2key(line, delim):
        return delim.join(map(str, line))
    for item in iterable:
        key = line2key([item[col] for col in keycols], delim)
        grouped_data.setdefault(key, {})
        for col in sumcols:
            grouped_data[key].setdefault(col, 0)
            grouped_data[key][col] += item[col]
        if to_count:
            grouped_data[key].setdefault("count", 0)
            grouped_data[key]["count"] += 1

    if to_count: sumcols += ["count"]
    header = keycols + sumcols
    for key,value in grouped_data.items():
        values = key.split(delim) + [value[col] for col in sumcols]
        yield dict([(k, v) for k,v in zip(header, values)])

## デバッグ系
@Pipe
def debug_print(iterable, colnames=None):
    """デバッグのために，行全体または一列のみを出力する"""
    for item in iterable:
        if colnames is None: print(item)
        else: print([item[colname] for colname in colnames])
        yield item

@Pipe
def show_progress(iterable, span=100000):
    """
    進捗どうですか
    ここを通過したデータ数をspanごとに標準エラー出力に表示
    """
    for i, item in enumerate(iterable):
        if i%span == 0:
            sys.stderr.write(str(i)+"\n")
        yield item

## 出力系
@Pipe
def cut(iterable, columns):
    """任意のカラムだけを残す．順序も保持されるようになるので，出力前の整形に用いることを想定

    ```
    (
        [{"name": "Alice", "sex": "F", "score": 90},
         {"name": "Bob", "sex": "M", "score": 20},
         {"name": "Chris", "sex": "M", "score": 30}]
        | juicypipe.cut(["sex", "score"])
        | juicypipe.csvprint
    )
    ```
    """
    return (collections.OrderedDict([(key, item[key]) for key in columns]) for item in iterable)


@Pipe
def csvprint(iterable, columns=None, delim="\t", to_yield_colname=False):
    """csv出力"""
    for item in iterable:
        if columns is None: columns = item.keys()
        if to_yield_colname:
            print(delim.join(columns))
            to_yield_colname = False
        print(delim.join(map(str, [item[key] for key in columns])))

@Pipe
def sqlite3_store(iterable, filename="pipestore.sqlite3", tablename="pipestore"):
    """sqlite3のデータベースとして出力する"""
    # init
    # 一行目を利用してテーブルのスキーマを得る
    first_row = iterable.next()
    connection = sqlite3.connect(filename)
    connection.execute("create table %s (%s)" % (tablename, ",".join(first_row.keys())))

    # store
    connection.execute('insert into %s values (%s)' % (tablename, ",".join(map(lambda x: '"%s"'%x, first_row.values()))))
    for item in iterable:
        connection.execute('insert into %s values (%s)' % (tablename, ",".join(map(lambda x: '"%s"'%x, item.values()))))

    connection.commit()
    connection.close()

@Pipe
def excelwrite(iterable, filename="pipestore.xlsx", sheetname="pipestore", mode="a", to_yield_colname=True, to_chain=False, col_offset=0, row_offset=0):
    """excelにしゅつりょくする

    :param mode: aなら存在するファイルに追記, wなら新規
    :param sheetname: 出力するシート名. 存在しているシート名を指定したら上書きされるので注意
    :param to_yield_colname: シートの一行目にカラム名を出力するか否か
    :type to_yield_colname: Bool
    :param to_chain: この後にpipeを繋げる場合はTrueにすること. ただしメモリを食う
    :type to_chain: Bool
    :param col_offset: エクセルに追記させる時の開始列の位置
    :param row_offset: エクセルに追記させる時の開始行の位置
    """
    if mode == "a" and os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.get_sheet_by_name(sheetname)
        if sheet is not None:
            # 存在してたら一旦消す
            workbook.remove_sheet(sheet)
        sheet = workbook.create_sheet()
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
    sheet.title = sheetname
    row = row_offset
    chain_gen = []
    for item in iterable:
        if to_yield_colname:
            for col, name in enumerate(item.keys()):
                sheet.cell(row=row+1, column=col+1+col_offset).value = name
            row += 1
            to_yield_colname = False
        for col, value in enumerate(item.values()):
            sheet.cell(row=row+1, column=col+1+col_offset).value = value
        row += 1
        if to_chain:
            chain_gen.append(item)
    workbook.save(filename)
    return chain_gen

if __name__ == '__main__':
    import doctest
    doctest.testmod()
