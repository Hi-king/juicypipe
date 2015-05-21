#-*- coding: utf-8 -*
import unittest
import sys
import os
script_path = os.path.dirname(__file__)
script_path = script_path if len(script_path) else "."
sys.path.insert(0, script_path+"/../")

import juicypipe
import pipe
import sqlite3
import mox
import openpyxl
from StringIO import StringIO

class JuicypipeTest(unittest.TestCase):
    def test_vertical_map_int(self):
        """vertical_mapで型変換"""
        result = (
            [{"val": "1"}, {"val": "2"}]
            | juicypipe.vertical_map("val", int)
            | pipe.as_list
        )
        self.assertEqual(result, [{"val": 1}, {"val": 2}])

    def test_vertical_map_lambda(self):
        """vertical_mapでλ演算"""
        result = (
            [{"val": "1"}, {"val": "2"}]
            | juicypipe.vertical_map("val", int)
            | juicypipe.vertical_map("val", lambda x: x**3)
            | pipe.as_list
        )
        self.assertEqual(result, [{"val": 1}, {"val": 8}])

    def test_vertical_map_lambda_newcolname(self):
        """vertical_mapでλ演算して新カラム追加"""
        result = (
            [{"val": "1"}, {"val": "2"}]
            | juicypipe.vertical_map("val", int)
            | juicypipe.vertical_map(
                colname="val", 
                func=lambda x: x**3,
                newcolname="pow")
            | pipe.as_list
        )
        self.assertEqual(result, [
                {"val": 1, "pow": 1},
                {"val": 2, "pow": 8}])

    def test_pipe_round(self):
        """小数点第一位で丸め"""
        result = (
            [{"val": 2.06}, {"val": 2.148}]
            | juicypipe.pipe_round("val", 1)
            | pipe.as_list
        )
        self.assertEqual(result, [{"val": 2.1}, {"val": 2.1}])

    def test_topivot(self):
        """カラムをグループ化してピボット形式にする"""
        result = (
            [{"name": "Alice", "sex": "F", "score": 90}, 
             {"name": "Bob", "sex": "M", "score": 20}, 
             {"name": "Chris", "sex": "M", "score": 30}]
            | juicypipe.topivot(
                keycols=["sex"],
                sumcols=["score"]
                )
            | pipe.as_list
        )
        self.assertEqual(
            result,
            [ {"sex": "M", "score": 50},
              {"sex": "F", "score": 90}]
        )

    def test_topivot_count(self):
        """カラムをグループ化して該当キー数を数える"""
        result = (
            [{"name": "Alice", "sex": "F", "score": 90}, 
             {"name": "Bob", "sex": "M", "score": 20}, 
             {"name": "Chris", "sex": "M", "score": 30}]
            | juicypipe.topivot(
                keycols=["sex"],
                sumcols=["score"],
                to_count=True
                )
            | pipe.as_list
        )
        self.assertEqual(
            result,
            [ {"sex": "M", "score": 50, "count": 2},
              {"sex": "F", "score": 90, "count": 1}]
        )

    def test_debug_print(self):
        """入ってきた一行のログをそのまま出力"""
        out = StringIO()
        original_stdout = sys.stdout
        sys.stdout = out
        (
            [{"name": "Alice", "sex": "F"},
             {"name": "Bob", "sex": "M"}]
            | juicypipe.debug_print
            | pipe.as_list
        )
        self.assertEqual(out.getvalue(), "{'name': 'Alice', 'sex': 'F'}\n{'name': 'Bob', 'sex': 'M'}\n")
        sys.stdout = original_stdout

    def test_debug_print_eachcol(self):
        """入ってきた一行のログから一列出力"""
        out = StringIO()
        original_stdout = sys.stdout
        sys.stdout = out
        (
            [{"name": "Alice", "val": 2.06},
             {"name": "Bob", "val": 2.148}]
            | juicypipe.debug_print(["name"])
            | pipe.as_list
        )
        self.assertEqual(out.getvalue(), "['Alice']\n['Bob']\n")
        sys.stdout = original_stdout

    def test_cut_and_order(self):
        """順序つきの部分辞書にされる"""
        result1 = (
            [{"name": "Alice", "sex": "F", "score": 90},
             {"name": "Bob", "sex": "M", "score": 20},
             {"name": "Chris", "sex": "M", "score": 30}]
            | juicypipe.cut(["sex", "score"])
            | pipe.as_list
        )
        self.assertListEqual(
            result1[0].values(),
            ["F", 90]
        )
        result2 = (
            [{"name": "Alice", "sex": "F", "score": 90},
             {"name": "Bob", "sex": "M", "score": 20},
             {"name": "Chris", "sex": "M", "score": 30}]
            | juicypipe.cut(["score", "sex"])
            | pipe.as_list
        )
        self.assertListEqual(
            result2[0].values(),
            [90, "F"]
        )

    def test_csvprint_schema(self):
        """スキーマ付きcsvとして出力"""
        out = StringIO()
        original_stdout = sys.stdout
        sys.stdout = out
        (
            [{"name": "Alice", "val": 2.06},
             {"name": "Bob", "val": 2.148}]
            | juicypipe.csvprint(delim="\t")
        )
        self.assertEqual(out.getvalue(), "Alice\t2.06\nBob\t2.148\n")
        sys.stdout = original_stdout

    def test_show_progress(self):
        """進捗状況が出力され，stdoutとstderrが分けられる"""
        test_stdout, test_stderr = StringIO(), StringIO()
        original_stdout, original_stderr = sys.stderr, sys.stdout
        sys.stdout, sys.stderr = test_stdout, test_stderr
        (
            [{"num": "A"},
             {"num": "B"},
             {"num": "C"},
             {"num": "D"},
             {"num": "E"},
             {"num": "F"},
             {"num": "G"},
             {"num": "H"},
             {"num": "I"},
             {"num": "J"}]
            | juicypipe.show_progress(2)
            | juicypipe.csvprint
        )
        self.assertEqual(test_stderr.getvalue(), "0\n2\n4\n6\n8\n")
        self.assertEqual(test_stdout.getvalue(), "A\nB\nC\nD\nE\nF\nG\nH\nI\nJ\n")
        sys.stdout = original_stdout

    def test_csvprint(self):
        """csvとして出力"""
        out = StringIO()
        original_stdout = sys.stdout
        sys.stdout = out
        (
            [{"name": "Alice", "val": 2.06},
             {"name": "Bob", "val": 2.148}]
            | juicypipe.csvprint(
                delim="\t",
                to_yield_colname=True
            )
        )
        self.assertEqual(out.getvalue(), "name\tval\nAlice\t2.06\nBob\t2.148\n")
        sys.stdout = original_stdout

    def test_sqlite3_store(self):
        """オンメモリsqlite3に入れる"""
        mock = mox.Mox()
        connection_mock = mock.CreateMock(sqlite3.Connection)
        mock.StubOutWithMock(sqlite3, 'connect')
        sqlite3.connect(mox.IsA(str)).AndReturn(connection_mock)
        connection_mock.execute('create table pipestoretest (name)')
        connection_mock.execute('insert into pipestoretest values ("Alice")')
        connection_mock.execute('insert into pipestoretest values ("Bob")')
        connection_mock.execute('insert into pipestoretest values ("")')
        connection_mock.commit()
        connection_mock.close()

        mock.ReplayAll()
        (
            iter([{"name": "Alice"},
                 {"name": "Bob"},
                 {"name": ""}])
            | juicypipe.sqlite3_store(
                filename=":memory:",
                tablename="pipestoretest"
            )
        )
        mock.VerifyAll()
        mock.UnsetStubs()

    def test_excelwrite(self):
        """excelに出力する

        新規作成モードで突っ込んだ時に, 一行目がカラム名になっていて, 以降にデータがちゃんと突っ込まれてることの確認
        また，シート名も指定した値になっていることを確認
        """
        mock = mox.Mox()
        dummy_workbook = openpyxl.Workbook()
        dummy_worksheet = dummy_workbook.active
        mock.StubOutWithMock(openpyxl.Workbook, 'active')
        openpyxl.Workbook.active = dummy_worksheet
        mock.StubOutWithMock(openpyxl.Workbook, 'save')
        openpyxl.Workbook.save(mox.IsA(str)).AndReturn(None)
        mock.ReplayAll()
        (
            [{"name": "Alice"},
            {"name": "Bob"},
            {"name": ""}]
            | juicypipe.excelwrite(sheetname="fakesheet")
        )
        mock.VerifyAll()
        mock.UnsetStubs()
        self.assertEqual(dummy_workbook.active.title, "fakesheet")
        self.assertEqual(dummy_workbook.active.cell("A1").value, "name")
        self.assertEqual(dummy_workbook.active.cell("A2").value, "Alice")
